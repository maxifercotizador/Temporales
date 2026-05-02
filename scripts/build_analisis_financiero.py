#!/usr/bin/env python3
"""
build_analisis_financiero.py
-----------------------------------------------------------------
Regenera analisis_financiero_datos.js mergeando los archivos
mensuales que viven en Archivos_Anal.Fiinan/YYYY-MM/.

Por cada mes, parsea los archivos que estén presentes:

  facturacion.xlsx        BS Gestion: comprobantes emitidos
  gastos_excel.xlsx       Excel propio del user (3 secciones)
  gastos_bs.xlsx          BS Gestion: gastos (comisiones reales)
  extracto_galicia.xlsx   Home banking Galicia
  resumen_santander.pdf   Resumen Santander del mes
  cobranzas.txt|png|jpg   Total cobrado del mes
  saldos.txt|png|jpg      Saldos al cierre del mes

Lo que ya estaba cargado en el JS (meses anteriores) se preserva
intacto. Solo se sobrescribe el mes para el cual hay archivos.

Uso:
  python3 scripts/build_analisis_financiero.py

Lo dispara la GitHub Action cuando alguien sube archivos a
Archivos_Anal.Fiinan/. El JS resultante se commitea automatico.
"""

import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
ARCHIVOS_DIR = ROOT / "Archivos_Anal.Fiinan"
JS_OUT = ROOT / "analisis_financiero_datos.js"

# Configuracion de proveedores/vendedores conocidos
VENDEDOR_PAPA = "Gordillo, Victor"
PAPA_CLIENTE_CODIGO = "00669"  # Maxifer Victor Distribuidor
TIPOS_NC = {"NCA", "NCB", "NCC", "NCM", "NC", "NCVA", "NCVB"}

# Tarjetas: matcheo por últimos dígitos (más confiable que palabras clave).
# El concepto en el extracto suele incluir los últimos 4 dígitos.
TARJETAS_DIGITOS = {
    "Tarjeta Galicia - VISA":      ["2884"],
    "Tarjeta Galicia - AMEX":      ["0793"],
    "Tarjeta Galicia - BUSINESS":  ["9091"],
    "Tarjeta Galicia + VISA":      ["3394"],
    "Tarjeta Galicia + MASTER":    ["6770"],
    "Tarjeta Santander Rio - VISA":["2857"],
    "Tarjeta Santander Rio - AMEX":["62044", "2044"],
    "Tarjeta ICBC":                ["7406"],
}
# Patrones de palabras clave como fallback si el extracto no tiene los dígitos
TARJETAS_PALABRAS = {
    "Tarjeta Galicia - VISA":      ["pago tarjeta visa", "pago visa galicia"],
    "Tarjeta Galicia - AMEX":      ["pago tarjeta amex galicia", "amex galicia"],
    "Tarjeta Galicia - BUSINESS":  ["business"],
    "Tarjeta Galicia + VISA":      ["plus visa", "+visa", "+ visa"],
    "Tarjeta Galicia + MASTER":    ["plus master", "+master", "+ master"],
    "Tarjeta Santander Rio - VISA":["pago tarjeta visa", "pago visa", "visa santander"],
    "Tarjeta Santander Rio - AMEX":["amex", "american express"],
    "Tarjeta ICBC":                ["icbc"],
}


# =====================================================================
# Helpers
# =====================================================================

def log(*args):
    print("[build]", *args, file=sys.stderr)


def s(v):
    if v is None:
        return ""
    return str(v).strip()


def num(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    txt = str(v).replace(",", ".").replace("$", "").strip()
    try:
        return float(txt)
    except ValueError:
        return 0


def parse_date(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(v, fmt)
            except ValueError:
                pass
    return None


def load_existing_js():
    if not JS_OUT.exists():
        return {
            "meses": [],
            "facturacion_por_mes": {},
            "gastos_por_mes": {},
            "volumen_por_mes": {},
            "gastos_lista": [],
            "fact_lista": [],
            "top_conceptos": {},
            "cobranzas_2026": {"total_cobrado": 0, "total_cobranzas": 0,
                               "por_mes": {}, "cobrado_papa_2026": 0,
                               "por_vendedor": []},
            "saldos_bancarios": [],
            "prestamos": [],
            "tarjetas_pagos": {},
            "problemas_tumini": {"monto": 0, "facturado_2026": 0},
        }
    txt = JS_OUT.read_text(encoding="utf-8")
    txt = txt.replace("window.DATOS = ", "").rstrip().rstrip(";")
    return json.loads(txt)


def write_js(data):
    payload = "window.DATOS = " + json.dumps(data, ensure_ascii=False) + ";\n"
    JS_OUT.write_text(payload, encoding="utf-8")


def remove_month_from_lista(lista, ano_mes):
    return [x for x in lista if x.get("AnoMes") != ano_mes]


# =====================================================================
# Parsers de archivos
# =====================================================================

def parse_facturacion_xlsx(path, ano_mes):
    """Parsea facturacion.xlsx exportado de BS Gestion.
    Devuelve: (resumen_mes, fact_entries, volumen)
    """
    try:
        import openpyxl
    except ImportError:
        log(f"  facturacion: openpyxl no disponible")
        return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        log(f"  facturacion: error abriendo {path.name}: {e}")
        return None

    # Detectar headers (busca primera fila con columnas tipicas)
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, r in enumerate(rows[:10]):
        cells = [s(c).lower() for c in r]
        if any("fecha" in c for c in cells) and any("total" in c or "importe" in c for c in cells):
            header_row = i
            break
    if header_row is None:
        log(f"  facturacion: no encuentro headers en {path.name}")
        return None

    headers = [s(c).lower() for c in rows[header_row]]

    def find_col(*keywords):
        for i, h in enumerate(headers):
            if all(k in h for k in keywords):
                return i
        return None

    col_fecha = find_col("fecha")
    col_cliente = find_col("cliente") or find_col("razon")
    col_vendedor = find_col("vendedor")
    col_tipo = find_col("tipo")
    col_nro = find_col("numero") or find_col("nro") or find_col("comprobante")
    col_total = find_col("total") or find_col("importe")
    col_items = find_col("items") or find_col("cantidad")

    if col_fecha is None or col_total is None:
        log(f"  facturacion: faltan columnas Fecha o Total en {path.name}")
        return None

    fact_entries = []
    total_bruto = 0
    ventas_papa = 0
    notas_credito = 0
    neto_maxi = 0
    clientes = set()
    items_total = 0
    pedidos = 0

    for r in rows[header_row + 1:]:
        f = parse_date(r[col_fecha]) if col_fecha < len(r) else None
        if not f:
            continue
        if f.strftime("%Y-%m") != ano_mes:
            continue
        cliente = s(r[col_cliente]) if col_cliente is not None and col_cliente < len(r) else ""
        vendedor = s(r[col_vendedor]) if col_vendedor is not None and col_vendedor < len(r) else ""
        tipo = s(r[col_tipo]).upper() if col_tipo is not None and col_tipo < len(r) else ""
        nro = s(r[col_nro]) if col_nro is not None and col_nro < len(r) else ""
        total = num(r[col_total]) if col_total < len(r) else 0
        items = int(num(r[col_items])) if col_items is not None and col_items < len(r) else 0

        es_nc = any(t in tipo for t in TIPOS_NC)
        # Detectar ventas papá: por vendedor explícito O por código de cliente "(00669)"
        es_papa = (vendedor == VENDEDOR_PAPA) or (PAPA_CLIENTE_CODIGO and PAPA_CLIENTE_CODIGO in cliente)

        if es_nc:
            notas_credito += -abs(total) if total > 0 else total
        else:
            total_bruto += total
            if es_papa:
                ventas_papa += total
            else:
                neto_maxi += total

        fact_entries.append({
            "Fecha": f.strftime("%Y-%m-%d"),
            "Cliente": cliente,
            "Vendedor": vendedor,
            "Tipo": tipo,
            "NroComp": nro,
            "AnoMes": ano_mes,
            "Total": total,
            "Items": items,
        })

        if not es_nc:
            pedidos += 1
            clientes.add(cliente)
            items_total += items

    resumen = {
        "total_bruto": round(total_bruto, 2),
        "ventas_papa": round(ventas_papa, 2),
        "notas_credito": round(notas_credito, 2),
        "neto_maxi": round(total_bruto - ventas_papa + notas_credito, 2),
    }
    volumen = {
        "pedidos": pedidos,
        "items": items_total,
        "clientes_unicos": len(clientes),
        "monto_promedio_pedido": round(total_bruto / pedidos, 2) if pedidos else 0,
    }
    return resumen, fact_entries, volumen


def parse_gastos_excel_xlsx(path, ano_mes):
    """Parsea el Excel personal del user. Soporta 2 formatos:

    A) Formato flat (4 columnas):
         Fecha | Concepto | Importe | Tipo
       donde Tipo ∈ {Sueldos, GFijos, Mercaderia, GVarios}

    B) Formato 3-secciones (legacy):
         cols 4-8:   Fecha | Concepto | Monto | Monto USD | Pague?  (GFijos)
         cols 10-13: Fecha | Proveedor | Facturado? | Importe        (Mercaderia)
         cols 15-18: Fecha | Concepto | Facturado? | Importe         (GVarios)

    Devuelve: (resumen_categorias, gastos_lista, top_conceptos)
    """
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        log(f"  gastos_excel: error abriendo: {e}")
        return None

    # Buscar la hoja
    ws = None
    for name in wb.sheetnames:
        if "real" in name.lower() or "gasto" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    # Detectar formato leyendo headers en filas 1-3
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    headers_str = " | ".join(s(c).lower() for r in rows for c in r if c)
    is_flat = ("tipo" in headers_str and "concepto" in headers_str and "importe" in headers_str)

    fijos = mercaderia = varios = sueldos = 0
    lista = []
    conceptos_count = Counter()

    if is_flat:
        # Detectar header row
        header_row = None
        for i, r in enumerate(rows):
            cells = [s(c).lower() for c in r if c]
            if "fecha" in cells and "concepto" in cells and "tipo" in cells:
                header_row = i
                break
        if header_row is None:
            header_row = 0
        headers = [s(c).lower() for c in rows[header_row]]
        col_fecha = next((i for i, h in enumerate(headers) if h == "fecha"), None)
        col_concepto = next((i for i, h in enumerate(headers) if h == "concepto"), None)
        col_importe = next((i for i, h in enumerate(headers) if h in ("importe", "monto")), None)
        col_tipo = next((i for i, h in enumerate(headers) if h == "tipo"), None)

        if col_fecha is None or col_concepto is None or col_importe is None or col_tipo is None:
            log(f"  gastos_excel (flat): faltan columnas")
            return None

        for r in ws.iter_rows(min_row=header_row + 2, values_only=True):
            if len(r) <= max(col_fecha, col_concepto, col_importe, col_tipo):
                continue
            f = parse_date(r[col_fecha])
            if not f or f.strftime("%Y-%m") != ano_mes:
                continue
            conc = s(r[col_concepto])
            monto = num(r[col_importe])
            tipo_raw = s(r[col_tipo])
            if not conc or monto <= 0:
                continue

            tipo = tipo_raw  # mantener como vino
            t_low = tipo_raw.lower()
            if "sueldo" in t_low:
                sueldos += monto
                tipo = "Sueldos"
            elif "merc" in t_low:
                mercaderia += monto
                tipo = "Mercaderia"
            elif "fijo" in t_low or "gfijo" in t_low:
                fijos += monto
                tipo = "GFijos"
            elif "vario" in t_low or "gvario" in t_low:
                varios += monto
                tipo = "GVarios"
            else:
                varios += monto  # default
                tipo = "GVarios"

            lista.append({
                "Fecha": f.strftime("%Y-%m-%d"),
                "Concepto": conc, "Importe": monto,
                "Tipo": tipo, "AnoMes": ano_mes,
            })
            conceptos_count[conc] += 1

    else:
        # Formato 3-secciones (legacy)
        for r in ws.iter_rows(min_row=3, values_only=True):
            if not r:
                continue
            # Sección 1: GFijos (cols 4-8)
            if len(r) > 6:
                f = parse_date(r[4])
                conc = s(r[5])
                monto = num(r[6])
                if f and f.strftime("%Y-%m") == ano_mes and conc and monto > 0:
                    tipo = "Sueldos" if "sueldo" in conc.lower() else "GFijos"
                    if tipo == "Sueldos":
                        sueldos += monto
                    else:
                        fijos += monto
                    lista.append({
                        "Fecha": f.strftime("%Y-%m-%d"),
                        "Concepto": conc, "Importe": monto,
                        "Tipo": tipo, "AnoMes": ano_mes,
                    })
                    conceptos_count[conc] += 1

            # Sección 2: Mercaderia (cols 10-13)
            if len(r) > 13:
                f = parse_date(r[10])
                prov = s(r[11])
                imp = num(r[13])
                if f and f.strftime("%Y-%m") == ano_mes and prov and imp > 0:
                    mercaderia += imp
                    lista.append({
                        "Fecha": f.strftime("%Y-%m-%d"),
                        "Concepto": prov, "Importe": imp,
                        "Tipo": "Mercaderia", "AnoMes": ano_mes,
                    })

            # Sección 3: GVarios (cols 15-18)
            if len(r) > 18:
                f = parse_date(r[15])
                conc = s(r[16])
                imp = num(r[18])
                if f and f.strftime("%Y-%m") == ano_mes and conc and imp > 0:
                    varios += imp
                    lista.append({
                        "Fecha": f.strftime("%Y-%m-%d"),
                        "Concepto": conc, "Importe": imp,
                        "Tipo": "GVarios", "AnoMes": ano_mes,
                    })

    resumen = {
        "mercaderia": round(mercaderia, 2),
        "fijos": round(fijos, 2),
        "varios": round(varios, 2),
        "sueldos": round(sueldos, 2),
    }
    top = [{"concepto": c, "veces": n} for c, n in conceptos_count.most_common(10)]
    return resumen, lista, top


def parse_gastos_bs_xlsx(path, ano_mes):
    """BS Gestion gastos del mes -> comisiones_reales (sumadas)."""
    try:
        import openpyxl
    except ImportError:
        return 0
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        log(f"  gastos_bs: error abriendo: {e}")
        return 0

    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, r in enumerate(rows[:10]):
        cells = [s(c).lower() for c in r]
        if any("concepto" in c or "detalle" in c or "referencia" in c or "descripcion" in c for c in cells) and \
           any("importe" in c or "total" in c or "monto" in c for c in cells):
            header_row = i
            break
    if header_row is None:
        return 0

    headers = [s(c).lower() for c in rows[header_row]]
    col_fecha = next((i for i, h in enumerate(headers) if "fecha" in h), None)
    # Concepto puede estar en "concepto", "detalle", "descripcion" o "referencia"
    col_concepto = next((i for i, h in enumerate(headers)
                         if "concepto" in h or "detalle" in h or "descripcion" in h
                         or "referencia" in h), None)
    col_importe = next((i for i, h in enumerate(headers)
                        if "importe" in h or "total" in h or "monto" in h), None)

    if col_fecha is None or col_importe is None:
        return 0

    total_comisiones = 0
    for r in rows[header_row + 1:]:
        f = parse_date(r[col_fecha]) if col_fecha < len(r) else None
        if not f or f.strftime("%Y-%m") != ano_mes:
            continue
        # Buscar "comision" en CUALQUIER columna de texto (referencia/concepto/etc)
        text_all = " ".join(s(v) for v in r if isinstance(v, str)).lower()
        if "comisión" in text_all or "comision" in text_all:
            total_comisiones += num(r[col_importe])

    return round(total_comisiones, 2)


def _norm_txt(t):
    """Normaliza texto: minúsculas + sin acentos."""
    import unicodedata
    if not t: return ""
    nfkd = unicodedata.normalize('NFKD', str(t))
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).lower()


def parse_extracto_bancario(path, ano_mes, banco_filtro):
    """Parsea un extracto bancario. Detecta pagos de tarjetas por últimos dígitos.
    Soporta múltiples formatos:
    - Galicia: Fecha | Movimiento | Débito | Crédito | Saldo
    - Santander: Fecha | Descripción | Caja de Ahorro | Cuenta Corriente | Saldo
    """
    try:
        import openpyxl
    except ImportError:
        return {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        log(f"  extracto {banco_filtro}: error: {e}")
        return {}

    pagos = defaultdict(float)

    tarjetas_relevantes = {
        nombre: digitos for nombre, digitos in TARJETAS_DIGITOS.items()
        if banco_filtro.lower() in nombre.lower()
    }
    palabras_relevantes = {
        nombre: pals for nombre, pals in TARJETAS_PALABRAS.items()
        if banco_filtro.lower() in nombre.lower()
    }

    KEYS_CONCEPT = {"concepto","detalle","descripcion","movimiento","operacion","referencia"}
    KEYS_DEBE    = {"debe","debito","egreso","salida"}
    KEYS_IMPORTE = {"importe","monto","caja de ahorro","cuenta corriente","cta corriente","cta. corriente","ahorro"}
    KEYS_SALDO   = {"saldo"}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Detectar header — la fila debe tener "fecha" + algun "concepto/movimiento/detalle"
        header_row = None
        for i, r in enumerate(rows[:30]):
            cells = [_norm_txt(c) for c in r]
            has_fecha = any("fecha" in c for c in cells)
            has_concept = any(any(k in c for k in KEYS_CONCEPT) for c in cells)
            if has_fecha and has_concept:
                header_row = i
                break
        if header_row is None:
            continue

        norm_headers = [_norm_txt(c) for c in rows[header_row]]

        col_fecha = next((i for i, h in enumerate(norm_headers) if "fecha" in h), None)
        col_concepto = next((i for i, h in enumerate(norm_headers)
                             if any(k in h for k in KEYS_CONCEPT)), None)
        cols_debe = [i for i, h in enumerate(norm_headers)
                     if any(k in h for k in KEYS_DEBE)]
        cols_importe = [i for i, h in enumerate(norm_headers)
                        if any(k in h for k in KEYS_IMPORTE) and not any(k in h for k in KEYS_SALDO)]

        if col_fecha is None or col_concepto is None:
            continue

        for r in rows[header_row + 1:]:
            f = parse_date(r[col_fecha]) if col_fecha < len(r) else None
            if not f or f.strftime("%Y-%m") != ano_mes:
                continue
            conc = s(r[col_concepto]) if col_concepto < len(r) else ""
            conc_norm = _norm_txt(conc)

            # Importe: probar columnas debe primero, después columnas importe
            importe = 0
            for c in cols_debe:
                if c < len(r):
                    v = abs(num(r[c]))
                    if v > 0:
                        importe = v
                        break
            if importe == 0:
                for c in cols_importe:
                    if c < len(r):
                        v = abs(num(r[c]))
                        if v > 0:
                            importe = v
                            break
            if importe <= 0:
                continue

            # Solo cargos negativos (pagos): si hay col_debe (Galicia) usamos esa.
            # Para Santander los importes vienen con signo: si el original es negativo, es egreso.
            # Re-chequeamos signo en columnas importe (que también pueden tener positivos):
            sign_neg = False
            for c in cols_debe + cols_importe:
                if c < len(r):
                    raw = num(r[c])
                    if raw < 0:
                        sign_neg = True
                        break
            if not sign_neg and not cols_debe:
                # Para Santander si no detectamos negativo, no es pago
                continue

            # ESTRICTO: el concepto debe contener "TARJETA" para considerar pago de tarjeta
            # (evita falsos matches con CBUs que tengan los dígitos)
            if "tarjeta" not in conc_norm and "tarj." not in conc_norm:
                continue

            # Match por dígitos (más específico)
            matched = False
            for tarjeta, digitos in tarjetas_relevantes.items():
                if any(d in conc for d in digitos):
                    pagos[tarjeta] += importe
                    matched = True
                    break
            if matched:
                continue

            # Fallback: palabras clave (genéricas tipo "PAGO TARJETA VISA")
            # Si el extracto no diferencia (ej. "PAGO TARJETA VISA" sin dígitos),
            # no asignamos a tarjeta específica → la asignación correcta viene del Excel personal.
            # Solo asignamos si hay match único.
            candidatos = []
            for tarjeta, palabras in palabras_relevantes.items():
                if any(p in conc_norm for p in palabras):
                    candidatos.append(tarjeta)
            if len(candidatos) == 1:
                pagos[candidatos[0]] += importe

    return {k: round(v, 2) for k, v in pagos.items()}


def parse_extracto_galicia(path, ano_mes):
    return parse_extracto_bancario(path, ano_mes, 'galicia')


def parse_resumen_tarjeta_pdf(path, ano_mes, tarjeta_nombre):
    """Parsea un PDF de resumen de tarjeta. Extrae:
       - total: total a pagar del mes (busca lineas tipo "TOTAL"/"PAGO MINIMO"/"VENCIMIENTO")
       - transacciones: lista de {fecha, concepto, monto, cuotas}

    Es un parser GENERICO que se va a refinar cuando tengamos PDFs de muestra
    de cada banco (Galicia, Santander, ICBC tienen formatos distintos).
    """
    try:
        import pdfplumber
    except BaseException as e:
        log(f"  PDF: pdfplumber no disponible ({type(e).__name__}: {e}), skip {path.name}")
        return None
    try:
        with pdfplumber.open(path) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception as e:
        log(f"  PDF {path.name}: error abriendo: {e}")
        return None

    info = {"total": 0, "transacciones": []}

    # ===== Total a pagar (heurística genérica) =====
    # Busca el monto más alto que aparezca cerca de palabras clave de "total/vencimiento"
    candidatos = []
    for line in text.splitlines():
        up = line.upper()
        if ("TOTAL" in up or "PAGO MINIMO" in up or "VENCIMIENTO" in up or "SALDO ACTUAL" in up):
            for m in re.finditer(r"\$?\s*([\d.,]+)\s*$|\$\s*([\d.,]+)", line):
                txt = (m.group(1) or m.group(2) or "").replace(".", "").replace(",", ".")
                try:
                    v = float(txt)
                    if v > 100:
                        candidatos.append(v)
                except ValueError:
                    pass
    if candidatos:
        info["total"] = round(max(candidatos), 2)

    # ===== Transacciones (heurística genérica) =====
    # Busca líneas con patrón: fecha (dd/mm o dd-mm-yy) + descripción + monto al final
    yy_corto = ano_mes[2:4]
    mm = int(ano_mes.split("-")[1])
    for line in text.splitlines():
        # Patrón: dd/mm + concepto + monto
        m = re.match(r"\s*(\d{2})[/-](\d{2})(?:[/-](\d{2,4}))?\s+(.+?)\s+\$?\s*([\d.,]+)\s*$", line)
        if not m:
            continue
        d, mes, _, concepto, monto = m.groups()
        try:
            d = int(d); mes = int(mes)
            if mes < 1 or mes > 12 or d < 1 or d > 31:
                continue
            v = float(monto.replace(".", "").replace(",", "."))
            if v < 1 or v > 100_000_000:
                continue
        except ValueError:
            continue
        # Cuotas (texto tipo "C.01/12")
        cuotas = None
        m_cuotas = re.search(r"C\.?\s*(\d{1,2})\s*/\s*(\d{1,2})", concepto)
        if m_cuotas:
            cuotas = f"{m_cuotas.group(1)}/{m_cuotas.group(2)}"
        info["transacciones"].append({
            "fecha": f"{d:02d}/{mes:02d}",
            "concepto": concepto.strip()[:100],
            "monto": round(v, 2),
            "cuotas": cuotas,
        })

    return info


def parse_resumen_santander_pdf(path, ano_mes):
    # Compat: ahora redirigimos al parser genérico (Santander VISA único)
    info = parse_resumen_tarjeta_pdf(path, ano_mes, "Tarjeta Santander Rio - VISA")
    if info and info.get("total"):
        return {"Tarjeta Santander Rio - VISA": info["total"]}
    return {}


def parse_cobranzas_txt(path):
    """Parsea cobranzas.txt formato:
       total_cobrado: 9100000
       cantidad: 30
       cobrado_papa: 1200000
    """
    out = {}
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        if ":" not in line:
            continue
        k, v = line.split(":", 1)
        k = k.strip().lower()
        out[k] = num(v)
    return out


def parse_saldos_txt(path):
    """Parsea saldos.txt formato:
       galicia: 2529212
       santander: 7613
       usd: 44.62
    """
    out = {}
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        if ":" not in line:
            continue
        k, v = line.split(":", 1)
        out[k.strip().lower()] = num(v)
    return out


def parse_consumos_txt(path):
    """Parsea archivos _resumen_*_consumos.txt y extrae transacciones detalladas.
    Soporta dos formatos:
      A) Detalle línea-a-línea bajo 'DETALLE MOVIMIENTOS' o 'DETALLE CONSUMOS':
         '  29/04/2026  CONCEPTO              USD 100,00'  o  '  ... $ 12.345,67'
      B) Solo categorías agregadas (sin línea-a-línea) — devuelve [].

    Devuelve lista de dicts: [{fecha, concepto, monto, moneda, tarjeta_sub}, ...]
    Si hay subcuentas (Nadia/Nora/Victor/Maxi), las detecta por contexto.
    """
    text = path.read_text(encoding="utf-8", errors="ignore")
    lines = text.split("\n")
    txns = []
    current_subcuenta = None
    in_detail = False
    in_cuotas = False
    for line in lines:
        ls = line.strip()
        # Detectar bloque de detalle
        upper = ls.upper()
        if "DETALLE MOVIMIENTOS" in upper or "DETALLE CONSUMOS" in upper:
            in_detail = True
            continue
        if in_detail and ("CATEGORÍAS" in upper or "CATEGORIAS" in upper or
                         "TOTALES" in upper or "NOTA:" in upper.replace("Á", "A")):
            in_detail = False
            continue
        # Detectar subcuenta
        m_sub = re.match(r'\s*(?:VISA|AMEX)[- ]?(\d{4}|\w+\d+)\s*\(([^)]+)\)', ls, re.IGNORECASE)
        if m_sub:
            current_subcuenta = m_sub.group(2).strip()
        m_sub2 = re.match(r'^([A-Z]+)\s+(?:titular|adicional)\s+(\w+)', ls)
        if m_sub2:
            current_subcuenta = m_sub2.group(2)

        if not in_detail:
            continue

        # Patrón de transacción:
        # '  29/04/2026  CONCEPTO ...  USD 100,00'
        # '  09/10/2025 - dloel outlet abertu (cuota 7/12) - $40.749,63'
        m = re.match(
            r'^\s*(\d{1,2}/\d{1,2}/\d{2,4})\s*[-–]?\s*(.*?)\s+(USD|U\$S|\$)\s*(-?[\d\.,]+)\s*$',
            line.replace('\xa0', ' ')
        )
        if m:
            fecha, concepto, moneda, monto_str = m.groups()
            # Formato argentino: "53.495,00" o "100,00" → puntos como miles, coma decimal
            ms = monto_str.strip()
            if "," in ms:
                ms = ms.replace(".", "").replace(",", ".")
            try:
                monto = float(ms)
            except ValueError:
                monto = 0
            txns.append({
                "fecha": fecha,
                "concepto": concepto.strip(),
                "monto": monto,
                "moneda": "USD" if moneda.upper().startswith("U") else "ARS",
                "subcuenta": current_subcuenta or "",
            })
    return txns


def parse_clientes_saldos_txt(path):
    """Parsea clientes_saldos.txt (TSV con header de comentarios):
       Cod\tCliente\tNombre Fantasia\tVendedor Asociado\tSaldo\tFecha
    Devuelve lista de dicts y total + fecha más reciente.
    """
    deudores = []
    fecha_max = ""
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.rstrip()
        if not line or line.startswith("=") or line.startswith("Saldos") or \
           line.startswith("Origen") or line.startswith("Formato") or \
           line.startswith("Columnas"):
            continue
        parts = line.split("\t")
        if len(parts) < 5:
            continue
        try:
            saldo = num(parts[4])
        except Exception:
            continue
        fecha = parts[5].strip() if len(parts) >= 6 else ""
        if fecha and fecha > fecha_max:
            fecha_max = fecha
        deudores.append({
            "codigo": parts[0].strip(),
            "cliente": parts[1].strip(),
            "fantasia": parts[2].strip(),
            "vendedor": parts[3].strip(),
            "saldo": saldo,
            "fecha": fecha,
        })
    deudores.sort(key=lambda d: -d["saldo"])
    return {
        "lista": deudores,
        "total": round(sum(d["saldo"] for d in deudores), 2),
        "fecha": fecha_max,
    }


def parse_ventas_detalle_tsv(path):
    """Parsea _ventas_detalle_raw.tsv (export de BS Gestion paste).
    Header: Fecha\tTipo\tNroComprobante\tCliente\tProducto\tNumero\tCantidad\t$ Neto\tPrecioTotal\tVendedor
    Devuelve lista de tuplas (producto, numero, precio_total, fecha).
    """
    rows = []
    text = path.read_text(encoding="utf-8", errors="ignore")
    lines = text.split("\n")
    if not lines:
        return rows
    for line in lines[1:]:
        parts = line.split("\t")
        if len(parts) < 10:
            continue
        fecha, tipo, nrocom, cliente, producto, numero, cant, neto, preciototal, vendedor = parts[:10]
        if not fecha or fecha.lower().startswith("total") or not producto:
            continue
        try:
            pt = float(preciototal.replace(".", "").replace(",", "."))
        except ValueError:
            continue
        # Neto (sin IVA): columna "$ Neto"
        try:
            neto_v = float(str(neto).replace(".", "").replace(",", "."))
        except ValueError:
            neto_v = 0
        try:
            cant_v = float(str(cant).replace(".", "").replace(",", "."))
        except ValueError:
            cant_v = 0
        rows.append({
            "fecha": fecha, "producto": producto, "numero": numero,
            "precio_total": pt,    # con IVA (lo que figura en factura)
            "neto": neto_v,        # sin IVA — usar para cálculos contables
            "cantidad": cant_v,
            "cliente": cliente, "vendedor": vendedor,
        })
    return rows


def cargar_listas_maxifer():
    """Carga el archivo Listas Maxifer.xlsx (con columna Fabrica) si está disponible.
    Devuelve dicts para mapear (Producto, Descripcion) -> Fabrica.
    """
    import unicodedata
    # Buscar Listas Maxifer en orden:
    #   1. scripts/data/Listas_Maxifer.xlsx (versión committeada para CI)
    #   2. ../Presupuestador/Listas Maxifer.xlsx (entorno local del user)
    candidatos = [
        ROOT / "scripts" / "data" / "Listas_Maxifer.xlsx",
        ROOT / ".." / "Presupuestador" / "Listas Maxifer.xlsx",
    ]
    listas_path = next((p for p in candidatos if p.exists()), None)
    if listas_path is None:
        log(f"  victor: no encuentro Listas Maxifer en {[str(p) for p in candidatos]}")
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(listas_path, read_only=True, data_only=True)
        ws = wb["LISTAS Y BD"]
    except Exception as e:
        log(f"  victor: error abriendo Listas Maxifer: {e}")
        return None

    def strip_accents(s):
        return ''.join(c for c in unicodedata.normalize('NFKD', str(s or '')) if not unicodedata.combining(c))

    def norm(s):
        s = strip_accents(s).strip().upper().replace("'", "").replace('"', "")
        return ' '.join(s.split())

    # Headers: ID, Producto, Codigo BS GESTION, Precio, Numero, Descripción,
    #          Cant. Minima, Fabrica, Codigo, COSTO
    prod_pos = {}    # (producto, str(numero_pos)) -> {fabrica, costo}
    prod_desc = {}   # (producto, descripcion_norm) -> {fabrica, costo}
    prod_fabricas = {}  # producto -> Counter de fabricas
    prod_costos = {}    # producto -> [costos] (para promedio si no se encuentra exacto)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        producto = norm(row[1])
        numero = row[4]
        desc = norm(row[5])
        fabrica = str(row[7]).strip().upper() if row[7] else "?"
        costo = float(row[9]) if row[9] is not None and isinstance(row[9], (int, float)) else 0
        info = {"fabrica": fabrica, "costo": costo}
        if numero is not None:
            prod_pos[(producto, str(numero).strip())] = info
        if desc:
            prod_desc[(producto, desc)] = info
        prod_fabricas.setdefault(producto, Counter())[fabrica] += 1
        prod_costos.setdefault(producto, []).append(costo)
    return {"pos": prod_pos, "desc": prod_desc, "fabricas": prod_fabricas,
            "costos_avg": prod_costos, "norm": norm}


# Aliases de productos: nombre en ventas -> nombre en Listas Maxifer
PRODUCTO_ALIAS = {
    "CONECTORES PARA COMBUSTIBLE": "CONECTORES DE COMBUSTIBLE",
    "GAVETA CH": "GAVETAS",
    "GAVETA GR": "GAVETAS",
    "GAVETA ME": "GAVETAS",
    "PLASTICOS IMPORTADOS": "PLASTICOS",
    "TORNILLOS Y BULONES ESPECIALES": "TORNILLERIA ESPECIAL",
}


def lookup_fabrica_costo(v, listas):
    """Para una línea de venta, devuelve (fabrica, costo_unitario).
    Si no se puede determinar el producto, devuelve (None, 0).
    """
    norm = listas["norm"]
    pn = norm(v["producto"])
    pn = PRODUCTO_ALIAS.get(pn, pn)
    nn_desc = norm(v["numero"])
    info = listas["desc"].get((pn, nn_desc))
    if info is None:
        info = listas["pos"].get((pn, str(v["numero"]).strip()))
    if info is not None:
        return info["fabrica"], info["costo"]
    # Sin match exacto: usar fabrica mayoritaria de la categoría + costo promedio
    fs = listas["fabricas"].get(pn)
    if not fs:
        return None, 0
    fabrica = list(fs.keys())[0] if len(fs) == 1 else fs.most_common(1)[0][0]
    costos = listas["costos_avg"].get(pn, [])
    avg = sum(costos) / len(costos) if costos else 0
    return fabrica, avg


def es_cliente_victor_distribuidor(cliente):
    """True si el cliente es la distribuidora interna de Víctor."""
    if not cliente:
        return False
    c = cliente.lower()
    # "Maxifer Victor Distribuidor" o "Maxifer, Victor Distribuidor" (con o sin coma)
    return "maxifer" in c and "victor" in c and "distribuidor" in c


def es_vendedor_victor(vendedor):
    """True si el vendedor es Víctor (papá)."""
    if not vendedor:
        return False
    return vendedor.strip().lower().startswith("gordillo, victor") or \
           vendedor.strip().lower() == "gordillo victor"


def calcular_ventas_maxifer(ventas, listas):
    """Total + breakdown de ventas con Fabrica == MAXIFER que VENDIÓ Maxi.
    Devuelve dict: {total, por_categoria: [{categoria, monto, count}],
                    por_cliente: [{cliente, monto, count}],
                    items: [{fecha, producto, numero, cliente, cantidad, monto}]}
    Excluye ventas a 'Maxifer Víctor Distribuidor' (cross-sale interno).
    """
    if not listas or not ventas:
        return {"total": 0.0, "por_categoria": [], "por_cliente": [], "items": []}
    total = 0.0
    por_cat = defaultdict(lambda: {"total": 0.0, "count": 0})
    por_cli = defaultdict(lambda: {"total": 0.0, "count": 0})
    items = []
    norm = listas["norm"]
    for v in ventas:
        if es_cliente_victor_distribuidor(v.get("cliente", "")):
            continue
        fabrica, _ = lookup_fabrica_costo(v, listas)
        if fabrica != "MAXIFER":
            continue
        # NETO sin IVA — para que el Cierre con Víctor sea consistente con costo
        # (la columna COSTO de Listas Maxifer también es neto sin IVA)
        monto = v.get("neto", 0)
        total += monto
        cat = norm(v["producto"])
        por_cat[cat]["total"] += monto
        por_cat[cat]["count"] += 1
        cli = (v.get("cliente") or "").strip()
        por_cli[cli]["total"] += monto
        por_cli[cli]["count"] += 1
        items.append({
            "fecha": v.get("fecha", ""),
            "producto": v.get("producto", ""),
            "numero": v.get("numero", ""),
            "cliente": cli,
            "cantidad": v.get("cantidad", 0),
            "monto": round(monto, 2),
        })
    return {
        "total": round(total, 2),
        "por_categoria": sorted(
            [{"categoria": k, "monto": round(d["total"], 2), "count": d["count"]}
             for k, d in por_cat.items()],
            key=lambda x: -x["monto"]),
        "por_cliente": sorted(
            [{"cliente": k, "monto": round(d["total"], 2), "count": d["count"]}
             for k, d in por_cli.items()],
            key=lambda x: -x["monto"]),
        "items": items,
    }


def calcular_costo_no_fabrica_victor(ventas, listas):
    """Total + breakdown del COSTO de productos no-MAXIFER que pasaron por la red de Víctor.
    INCLUYE devoluciones (cantidad negativa) — restan al total porque son notas
    de crédito que anulan ventas previas.
    Para items sin costo cargado en Listas Maxifer, se ESTIMA usando el ratio
    promedio costo/retail del mismo proveedor (o 50% como fallback).
    Devuelve dict con total, breakdowns y items, marcando los estimados.
    """
    if not listas or not ventas:
        return {"total": 0.0, "por_categoria": [], "por_cliente": [],
                "por_fabrica": [], "items": [], "estimados_count": 0,
                "estimados_total": 0.0}
    norm = listas["norm"]

    # Primero: armar ratios costo/retail por fábrica usando items que SÍ tienen costo
    ratio_por_fabrica = {}
    suma_costo_fab = defaultdict(float)
    suma_retail_fab = defaultdict(float)
    for v in ventas:
        if not (es_cliente_victor_distribuidor(v.get("cliente", "")) or
                es_vendedor_victor(v.get("vendedor", ""))):
            continue
        fabrica, costo_unit = lookup_fabrica_costo(v, listas)
        if fabrica == "MAXIFER" or not fabrica:
            continue
        try:
            cant = float(str(v.get("cantidad", "0")).replace(",", "."))
        except Exception:
            continue
        if cant <= 0 or costo_unit <= 0:
            continue
        suma_costo_fab[fabrica] += cant * costo_unit
        suma_retail_fab[fabrica] += v.get("neto", 0)
    for fab, costo_acum in suma_costo_fab.items():
        retail_acum = suma_retail_fab[fab]
        if retail_acum > 0:
            ratio_por_fabrica[fab] = costo_acum / retail_acum
    # Ratio promedio global (fallback)
    total_costo_glob = sum(suma_costo_fab.values())
    total_retail_glob = sum(suma_retail_fab.values())
    ratio_global = (total_costo_glob / total_retail_glob) if total_retail_glob else 0.45

    total_costo = 0.0
    por_cat = defaultdict(lambda: {"total": 0.0, "count": 0})
    por_cli = defaultdict(lambda: {"total": 0.0, "count": 0})
    por_fab = defaultdict(lambda: {"total": 0.0, "count": 0})
    items = []
    estimados_count = 0
    estimados_total = 0.0
    for v in ventas:
        if not (es_cliente_victor_distribuidor(v.get("cliente", "")) or
                es_vendedor_victor(v.get("vendedor", ""))):
            continue
        fabrica, costo_unit = lookup_fabrica_costo(v, listas)
        if fabrica == "MAXIFER":
            continue
        try:
            cant = float(str(v.get("cantidad", "0")).replace(",", "."))
        except Exception:
            cant = 0
        retail_line = v.get("neto", 0)  # neto sin IVA, consistente con COSTO

        # Determinar costo del item:
        #  1. Si hay costo_unit definido → cant * costo_unit (incluye negativos por devoluciones)
        #  2. Si no hay costo_unit pero sí fábrica → estimar con ratio costo/retail de la fábrica
        #  3. Si tampoco hay fábrica → estimar con ratio global
        estimado = False
        if costo_unit > 0 and cant != 0:
            costo_total_item = cant * costo_unit
        else:
            ratio = ratio_por_fabrica.get(fabrica, ratio_global)
            costo_total_item = retail_line * ratio
            estimado = True
            estimados_count += 1
            estimados_total += costo_total_item

        if costo_total_item == 0:
            continue
        total_costo += costo_total_item
        cat = norm(v["producto"])
        por_cat[cat]["total"] += costo_total_item
        por_cat[cat]["count"] += 1
        cli = (v.get("cliente") or "").strip()
        por_cli[cli]["total"] += costo_total_item
        por_cli[cli]["count"] += 1
        fab_label = fabrica or "?"
        por_fab[fab_label]["total"] += costo_total_item
        por_fab[fab_label]["count"] += 1
        items.append({
            "fecha": v.get("fecha", ""),
            "producto": v.get("producto", ""),
            "numero": v.get("numero", ""),
            "cliente": cli,
            "vendedor": v.get("vendedor", ""),
            "fabrica": fab_label,
            "cantidad": cant,
            "costo_unit": round(costo_unit, 2),
            "monto": round(costo_total_item, 2),
            "estimado": estimado,
        })
    return {
        "total": round(total_costo, 2),
        "estimados_count": estimados_count,
        "estimados_total": round(estimados_total, 2),
        "ratio_por_fabrica": {f: round(r, 4) for f, r in ratio_por_fabrica.items()},
        "ratio_global": round(ratio_global, 4),
        "por_categoria": sorted(
            [{"categoria": k, "monto": round(d["total"], 2), "count": d["count"]}
             for k, d in por_cat.items()],
            key=lambda x: -x["monto"]),
        "por_cliente": sorted(
            [{"cliente": k, "monto": round(d["total"], 2), "count": d["count"]}
             for k, d in por_cli.items()],
            key=lambda x: -x["monto"]),
        "por_fabrica": sorted(
            [{"fabrica": k, "monto": round(d["total"], 2), "count": d["count"]}
             for k, d in por_fab.items()],
            key=lambda x: -x["monto"]),
        "items": items,
    }


# Materia prima fábrica: keywords que identifican proveedores/conceptos
MATERIA_PRIMA_KEYWORDS = [
    "alumina", "argon", "j l metales", "jl metales", "soldanex",
    "vische", "billordo", "cerminaro", "flejes", "varillas", "caños lumina",
    "caños de aluminio", "caños de alum",
    "aceros inoxidables", "union y fuerza", "soldaduras",
    "piso de goma", "pisos de goma", "gases y soldaduras",
]

# Sueldos fábrica: nombres
SUELDOS_FABRICA_KEYWORDS = ["nacho", "wilber", "lucas"]


def calcular_materia_prima_bs(bs_path, ano_mes):
    """Extrae materia prima desde gastos_bs.xlsx (tiene proveedores explícitos)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(bs_path, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        log(f"  victor: error abriendo gastos_bs: {e}")
        return 0.0, set()

    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, r in enumerate(rows[:10]):
        cells = [s(c).lower() for c in r]
        if any("fecha" in c for c in cells) and any("importe" in c or "total" in c for c in cells):
            header_row = i
            break
    if header_row is None:
        return 0.0, set()

    total = 0.0
    montos_vistos = set()
    for r in rows[header_row + 1:]:
        f = parse_date(r[0]) if r else None
        if not f or f.strftime("%Y-%m") != ano_mes:
            continue
        # Concatenar todo el texto del row para buscar keywords
        text_all = " ".join(s(v) for v in r if isinstance(v, str)).lower()
        if any(k in text_all for k in MATERIA_PRIMA_KEYWORDS):
            # Buscar el monto: última columna numérica
            importe = 0
            for v in reversed(r):
                if isinstance(v, (int, float)) and v != 0:
                    importe = abs(v)
                    break
            if importe > 0:
                total += importe
                montos_vistos.add(round(importe, 0))
    return round(total, 2), montos_vistos


def calcular_pusiste(gastos_lista, ano_mes, bs_path=None):
    """Suma sueldos fábrica + materia prima del mes (cruzando excel + bs sin duplicar)."""
    sueldos = 0.0
    materia_prima_bs = 0.0
    montos_bs = set()
    if bs_path:
        materia_prima_bs, montos_bs = calcular_materia_prima_bs(bs_path, ano_mes)

    materia_prima_excel = 0.0
    for g in gastos_lista:
        if g.get("AnoMes") != ano_mes:
            continue
        concepto = (g.get("Concepto") or "").lower()
        importe = abs(num(g.get("Importe")))
        if any(k in concepto for k in SUELDOS_FABRICA_KEYWORDS):
            sueldos += importe
            continue
        if any(k in concepto for k in MATERIA_PRIMA_KEYWORDS):
            # Si ya está en BS, no duplicar
            if round(importe, 0) in montos_bs:
                continue
            materia_prima_excel += importe
    materia_prima = round(materia_prima_bs + materia_prima_excel, 2)
    return round(sueldos, 2), materia_prima


# =====================================================================
# Procesar carpeta de un mes
# =====================================================================

def process_month(month_dir, data):
    name = month_dir.name  # "2026-04"
    if not re.match(r"\d{4}-\d{2}$", name):
        return False
    ano_mes = name

    # Recolectar todos los archivos: raíz del mes + subcarpetas (Excels, Galicia, Santander, ICBC)
    files = {}
    for f in month_dir.rglob("*"):
        if not f.is_file():
            continue
        if f.name.lower() in ("leeme.md", ".gitkeep"):
            continue
        # Key: tuple (subcarpeta_lower, filename_lower)
        rel_parent = f.parent.name.lower() if f.parent != month_dir else ""
        files[(rel_parent, f.name.lower())] = f

    if not files:
        return False

    def find(folder, name_lower):
        return files.get((folder.lower(), name_lower.lower()))

    log(f"Procesando {ano_mes} ({len(files)} archivos en {len(set(k[0] for k in files))} carpetas)...")
    cambios = []

    # 1. facturacion.xlsx (en Excels/)
    fpath = find("Excels", "facturacion.xlsx") or find("", "facturacion.xlsx")
    if fpath:
        r = parse_facturacion_xlsx(fpath, ano_mes)
        if r:
            resumen, entries, vol = r
            data["facturacion_por_mes"][ano_mes] = resumen
            data["volumen_por_mes"][ano_mes] = vol
            data["fact_lista"] = remove_month_from_lista(data.get("fact_lista", []), ano_mes) + entries
            cambios.append(f"facturacion ({len(entries)} comprobantes)")

    # 2. gastos_excel.xlsx (en Excels/)
    fpath = find("Excels", "gastos_excel.xlsx") or find("", "gastos_excel.xlsx")
    if fpath:
        r = parse_gastos_excel_xlsx(fpath, ano_mes)
        if r:
            resumen, lista, top = r
            existing = data["gastos_por_mes"].get(ano_mes, {})
            existing.update(resumen)
            existing["total"] = round(sum(resumen.values()), 2) + existing.get("comisiones_reales", 0)
            data["gastos_por_mes"][ano_mes] = existing
            data["gastos_lista"] = remove_month_from_lista(data.get("gastos_lista", []), ano_mes) + lista
            data["top_conceptos"][ano_mes] = top
            cambios.append(f"gastos_excel ({len(lista)} items)")

            # Detectar tarjetas en los items y actualizar tarjetas_pagos
            # Usa los nombres canónicos de TARJETAS_DIGITOS para matchear
            tarjetas_canon = list(TARJETAS_DIGITOS.keys())
            for item in lista:
                conc = item["Concepto"]
                conc_norm = _norm_txt(conc)
                # Match exacto si el concepto coincide con el nombre canónico (sin acentos)
                for tarjeta_canon in tarjetas_canon:
                    if _norm_txt(tarjeta_canon) == conc_norm:
                        data["tarjetas_pagos"].setdefault(tarjeta_canon, {})[ano_mes] = item["Importe"]
                        break

    # 3. gastos_bs.xlsx (en Excels/)
    fpath = find("Excels", "gastos_bs.xlsx") or find("", "gastos_bs.xlsx")
    if fpath:
        comisiones = parse_gastos_bs_xlsx(fpath, ano_mes)
        if comisiones:
            existing = data["gastos_por_mes"].get(ano_mes, {})
            existing["comisiones_reales"] = comisiones
            existing["total"] = round(
                existing.get("mercaderia", 0) + existing.get("fijos", 0) +
                existing.get("varios", 0) + existing.get("sueldos", 0) + comisiones, 2)
            data["gastos_por_mes"][ano_mes] = existing
            cambios.append(f"comisiones_reales ${comisiones:,.0f}")

    # 4a/b. Extractos bancarios (Galicia + Santander)
    # IMPORTANTE: gastos_excel.xlsx ya cargó las tarjetas con la asignación correcta
    # del user. El extracto solo COMPLEMENTA tarjetas que no estén ya cargadas
    # (porque el extracto suele decir "PAGO TARJETA VISA" sin dígitos y no permite
    # distinguir entre VISA-2884 y VISA-3394, etc).
    for nombre_arch, banco in [("extracto_galicia.xlsx", "galicia"),
                                ("extracto_santander.xlsx", "santander")]:
        fpath = find("Excels", nombre_arch) or find("", nombre_arch)
        if not fpath:
            continue
        pagos = parse_extracto_bancario(fpath, ano_mes, banco)
        sumados = 0
        for tarjeta, monto in pagos.items():
            existing = data["tarjetas_pagos"].get(tarjeta, {}).get(ano_mes)
            # Solo agregar si NO había valor previo o si era 0
            if existing is None or existing == 0:
                data["tarjetas_pagos"].setdefault(tarjeta, {})[ano_mes] = monto
                sumados += 1
        if sumados:
            cambios.append(f"{banco} tarjetas ({sumados} desde extracto)")

    # 5. PDFs de tarjetas (Galicia/, Santander/, ICBC/)
    pdf_specs = [
        ("Galicia",   "resumen_visa.pdf",         "Tarjeta Galicia - VISA"),
        ("Galicia",   "resumen_amex.pdf",         "Tarjeta Galicia - AMEX"),
        ("Galicia",   "resumen_business.pdf",     "Tarjeta Galicia - BUSINESS"),
        ("Galicia",   "resumen_plus_visa.pdf",    "Tarjeta Galicia + VISA"),
        ("Galicia",   "resumen_plus_master.pdf",  "Tarjeta Galicia + MASTER"),
        ("Santander", "resumen_visa.pdf",         "Tarjeta Santander Rio - VISA"),
        ("Santander", "resumen_amex.pdf",         "Tarjeta Santander Rio - AMEX"),
        ("ICBC",      "resumen_visa.pdf",         "Tarjeta ICBC"),
    ]
    pdfs_procesados = 0
    for folder, fname, tarjeta_nombre in pdf_specs:
        fpath = find(folder, fname)
        if not fpath:
            continue
        info = parse_resumen_tarjeta_pdf(fpath, ano_mes, tarjeta_nombre)
        if info and info.get("total"):
            data["tarjetas_pagos"].setdefault(tarjeta_nombre, {})[ano_mes] = info["total"]
            # Detalle de transacciones para análisis
            if info.get("transacciones"):
                data.setdefault("tarjetas_detalle", {}) \
                    .setdefault(tarjeta_nombre, {})[ano_mes] = info["transacciones"]
            pdfs_procesados += 1
    if pdfs_procesados:
        cambios.append(f"PDFs tarjeta ({pdfs_procesados})")

    # 5b. Fallback: archivos _resumen_*_consumos.txt (cuando los PDFs no parsean
    # o cuando hay datos manualmente cargados desde imágenes)
    consumos_specs = [
        ("Galicia",   "_resumen_visa_consumos.txt",     "Tarjeta Galicia - VISA"),
        ("Galicia",   "_resumen_amex_consumos.txt",     "Tarjeta Galicia - AMEX"),
        ("Galicia",   "_resumen_business_consumos.txt", "Tarjeta Galicia - BUSINESS"),
        ("Galicia",   "_resumen_plus_visa_consumos.txt","Tarjeta Galicia + VISA"),
        ("Galicia",   "_resumen_plus_master_consumos.txt","Tarjeta Galicia + MASTER"),
        ("Santander", "_resumen_visa_consumos.txt",     "Tarjeta Santander Rio - VISA"),
        ("Santander", "_resumen_amex_consumos.txt",     "Tarjeta Santander Rio - AMEX"),
        ("ICBC",      "_resumen_visa_consumos.txt",     "Tarjeta ICBC"),
    ]
    consumos_procesados = 0
    for folder, fname, tarjeta_nombre in consumos_specs:
        fpath = find(folder, fname)
        if not fpath:
            continue
        txns = parse_consumos_txt(fpath)
        if not txns:
            continue
        # Sobrescribir el detalle del mes (el .txt es la fuente de verdad reciente)
        data.setdefault("tarjetas_detalle", {}) \
            .setdefault(tarjeta_nombre, {})[ano_mes] = txns
        consumos_procesados += 1
    if consumos_procesados:
        cambios.append(f"consumos.txt ({consumos_procesados})")

    # 6. cobranzas.txt (en raíz del mes)
    fpath = find("", "cobranzas.txt")
    if fpath:
        c = parse_cobranzas_txt(fpath)
        if c.get("total_cobrado"):
            cob = data.get("cobranzas_2026", {})
            por_mes = cob.setdefault("por_mes", {})
            por_mes[ano_mes] = {
                "monto": int(c.get("total_cobrado", 0)),
                "cantidad": int(c.get("cantidad", 0)),
            }
            cob["total_cobrado"] = sum(v["monto"] for v in por_mes.values())
            cob["total_cobranzas"] = sum(v["cantidad"] for v in por_mes.values())
            data["cobranzas_2026"] = cob
            cambios.append(f"cobranzas ${c['total_cobrado']:,.0f}")

    # 7. saldos.txt (en raíz del mes)
    fpath = find("", "saldos.txt")
    if fpath:
        sal = parse_saldos_txt(fpath)
        if sal:
            yr, mo = map(int, ano_mes.split("-"))
            from calendar import monthrange
            last_day = monthrange(yr, mo)[1]
            fecha = f"{ano_mes}-{last_day:02d}"
            # Soporta 4 cuentas Galicia (principal, caja_ahorro, plus, caja_ahorro_plus)
            # más santander + usd. Compatible con formato viejo (galicia/santander/usd).
            galicia_total = (
                sal.get("galicia_principal", 0) +
                sal.get("galicia_caja_ahorro", 0) +
                sal.get("galicia_plus", 0) +
                sal.get("galicia_caja_ahorro_plus", 0) +
                sal.get("galicia", 0)  # backwards compat
            )
            entry = {
                "fecha": fecha,
                "galicia": int(galicia_total),
                "santander": int(sal.get("santander", 0)),
                "usd": sal.get("usd", 0),
            }
            entry["total"] = entry["galicia"] + entry["santander"]
            # Detalle de las 4 cuentas Galicia (si vienen)
            if any(k in sal for k in ("galicia_principal", "galicia_caja_ahorro",
                                      "galicia_plus", "galicia_caja_ahorro_plus")):
                entry["detalle"] = {
                    "galicia_principal": sal.get("galicia_principal", 0),
                    "galicia_caja_ahorro": sal.get("galicia_caja_ahorro", 0),
                    "galicia_plus": sal.get("galicia_plus", 0),
                    "galicia_caja_ahorro_plus": sal.get("galicia_caja_ahorro_plus", 0),
                }
            sb = data.get("saldos_bancarios", [])
            sb = [e for e in sb if e.get("fecha") != fecha]
            sb.append(entry)
            sb.sort(key=lambda e: e.get("fecha", ""))
            data["saldos_bancarios"] = sb
            cambios.append(f"saldos al {fecha}")

    # 8. clientes_saldos.txt -> deudores (último mes manda)
    fpath = find("", "clientes_saldos.txt")
    if fpath:
        d = parse_clientes_saldos_txt(fpath)
        # Solo sobrescribimos si este mes es más reciente que lo que ya tenemos
        prev = data.get("deudores", {})
        if d["fecha"] >= prev.get("fecha", ""):
            data["deudores"] = d
            cambios.append(f"deudores ({len(d['lista'])} clientes, ${d['total']:,.0f})")

    # 9. _ventas_detalle_raw.tsv -> Cierre con Víctor del mes.
    # Lógica del cálculo:
    #   PUSISTE = sueldos fábrica + materia prima fábrica + costo_no_fabrica_victor
    #     · costo_no_fabrica_victor = costo de productos NO-MAXIFER vendidos por
    #       Víctor (vendedor) o a su distribuidora (cliente). Es lo que vos pagaste
    #       a tus proveedores por mercadería que después él se quedó.
    #   RECIBISTE = ventas a precio retail de productos MAXIFER que vendiste vos
    #     (excluye ventas a 'Maxifer Víctor Distribuidor' = cross-sale interno).
    fpath = find("Excels", "_ventas_detalle_raw.tsv")
    if fpath:
        ventas = parse_ventas_detalle_tsv(fpath)
        listas = cargar_listas_maxifer()
        if ventas and listas:
            recibido = calcular_ventas_maxifer(ventas, listas)
            puesto_costo = calcular_costo_no_fabrica_victor(ventas, listas)
            bs_path = find("Excels", "gastos_bs.xlsx") or find("", "gastos_bs.xlsx")
            sueldos, materia_prima = calcular_pusiste(data.get("gastos_lista", []), ano_mes, bs_path)

            ventas_maxifer_total = recibido["total"]
            costo_no_fab_total = puesto_costo["total"]

            # Override manual: _victor_overrides.json en Excels/
            ov_path = find("Excels", "_victor_overrides.json")
            if ov_path:
                try:
                    ov = json.loads(Path(ov_path).read_text(encoding="utf-8"))
                    if "sueldos_fabrica" in ov:
                        sueldos = float(ov["sueldos_fabrica"])
                    if "materia_prima_extra" in ov:
                        materia_prima += float(ov["materia_prima_extra"])
                    if "costo_no_fabrica_extra" in ov:
                        costo_no_fab_total += float(ov["costo_no_fabrica_extra"])
                except Exception as e:
                    log(f"  victor: error leyendo overrides: {e}")
            pusiste = round(sueldos + materia_prima + costo_no_fab_total, 2)
            # Detectar mes parcial: si última fecha < día 28
            ultimas = sorted({v["fecha"] for v in ventas if v.get("fecha")})
            parcial = False
            if ultimas:
                ult = ultimas[-1]
                try:
                    dia = int(ult.split("/")[0])
                    parcial = dia < 28
                except Exception:
                    pass
            data.setdefault("cierre_victor", {}).setdefault("por_mes", {})[ano_mes] = {
                "sueldos": sueldos,
                "materia_prima": materia_prima,
                "costo_no_fabrica_victor": round(costo_no_fab_total, 2),
                "pusiste": pusiste,
                "ventas_maxifer": ventas_maxifer_total,
                "recibiste": ventas_maxifer_total,
                "parcial": parcial,
                # Detalle para que el dashboard pueda mostrar el desglose
                "detalle_recibido": {
                    "por_categoria": recibido["por_categoria"],
                    "por_cliente": recibido["por_cliente"],
                    # Top 200 items para que el JSON no explote
                    "items": sorted(recibido["items"], key=lambda x: -x["monto"])[:200],
                },
                "detalle_costo_victor": {
                    "por_categoria": puesto_costo["por_categoria"],
                    "por_cliente": puesto_costo["por_cliente"],
                    "por_fabrica": puesto_costo["por_fabrica"],
                    # Ordeno por |monto| desc para que devoluciones grandes (negativas)
                    # también aparezcan arriba
                    "items": sorted(puesto_costo["items"], key=lambda x: -abs(x["monto"]))[:300],
                    "estimados_count": puesto_costo.get("estimados_count", 0),
                    "estimados_total": puesto_costo.get("estimados_total", 0),
                    "ratio_por_fabrica": puesto_costo.get("ratio_por_fabrica", {}),
                    "ratio_global": puesto_costo.get("ratio_global", 0),
                },
            }
            cambios.append(f"victor: puso ${pusiste:,.0f} (sueldos+mp+costoVic) / recibió ${ventas_maxifer_total:,.0f}")

    if cambios and ano_mes not in data.get("meses", []):
        data.setdefault("meses", []).append(ano_mes)
        data["meses"].sort()

    if cambios:
        log(f"  {ano_mes}: " + ", ".join(cambios))
        return True
    return False


# =====================================================================
# Main
# =====================================================================

def main():
    if not ARCHIVOS_DIR.exists():
        log("No existe", ARCHIVOS_DIR)
        return 0

    data = load_existing_js()
    cambios_total = 0
    for month_dir in sorted(ARCHIVOS_DIR.iterdir()):
        if not month_dir.is_dir():
            continue
        if process_month(month_dir, data):
            cambios_total += 1

    if cambios_total:
        write_js(data)
        log(f"OK: {cambios_total} mes(es) actualizados → {JS_OUT.name}")
    else:
        log("Sin cambios — ningún mes tenía archivos para procesar")
    return 0


if __name__ == "__main__":
    sys.exit(main())
